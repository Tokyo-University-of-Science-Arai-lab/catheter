#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sam_recognition_eval の結果から、目視レビュー用のExcelレポートと
インスタンスごとのマスク重畳クロップ画像をまとめて作る。

前提: run_sam_eval.py が既に走っており、
  catheter/outputs/sam_recognition_eval/pred_masks_cache/{image_id}.npz
に予測マスクがキャッシュ済みであること(GPU再推論はしない、キャッシュを読むだけ)。

出力(新規のみ):
  catheter/outputs/sam_recognition_eval/catheter_recognition_report.xlsx
  catheter/outputs/sam_recognition_eval/mask_overlays/*.png   (インスタンスごとの拡大画像)
"""
from __future__ import annotations

import json
import csv
import re
import unicodedata
from pathlib import Path

import numpy as np
import cv2
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage

import sys
CATHETER_ROOT = Path(__file__).resolve().parents[1]  # .../catheter_test80/catheter
SRC = CATHETER_ROOT.parent / "catheter-100"
IMAGES_DIR = SRC / "images" / "default"
ANNOTATIONS_JSON = SRC / "annotations" / "instances_default.json"
MASTER_JSON = SRC / "master_catheter_20260216.json"

OCR_DIR = CATHETER_ROOT / "outputs" / "ocr"
REF_ASSIGNMENT_CSV = CATHETER_ROOT / "outputs" / "match_multi" / "assignment.csv"

EVAL_DIR = CATHETER_ROOT / "outputs" / "sam_recognition_eval"
MASK_CACHE_DIR = EVAL_DIR / "pred_masks_cache"
OVERLAY_CROP_DIR = EVAL_DIR / "mask_overlays"
OVERLAY_CROP_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(CATHETER_ROOT / "scripts"))
from compare_quad_fit import segmentation_to_mask, iou  # noqa: E402
from mask_rectify import min_area_rect_box  # noqa: E402
from match_eval import key_score  # noqa: E402
from scipy.optimize import linear_sum_assignment  # noqa: E402

IOU_MATCH_THRESHOLD = 0.5
MIN_REC_SCORE = 0.5
CROP_MARGIN = 25

# 4枚目の画像はユーザー指示により評価対象から除外する(4枚 x 20クエリ = 80件で評価)。
EXCLUDED_IMAGE_IDS = {4}


def safe_name(s: str, maxlen: int = 40) -> str:
    """display_name をファイル名に使える形へ(render_overlays_named.pyのsafe_nameを踏襲)。"""
    if not s:
        return "unknown"
    s = unicodedata.normalize("NFKC", s)
    for ch in ("®", "™", "©"):
        s = s.replace(ch, "")
    s = re.sub(r'[\\/:*?"<>|]', "", s)      # ファイル名に使えない文字
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return (s[:maxlen] or "unknown")


def rect_width_px(mask01: np.ndarray) -> float | None:
    box = min_area_rect_box(mask01)
    if box is None:
        return None
    w1 = float(np.linalg.norm(box[0] - box[1]))
    w2 = float(np.linalg.norm(box[1] - box[2]))
    return min(w1, w2)


def to_float_or_none(s) -> float | None:
    try:
        return float(str(s).strip())
    except (TypeError, ValueError):
        return None


def main() -> None:
    with open(ANNOTATIONS_JSON, "r", encoding="utf-8") as f:
        coco = json.load(f)
    images_by_id = {im["id"]: im for im in coco["images"]}
    anns_by_image: dict[int, list] = {}
    for ann in coco["annotations"]:
        anns_by_image.setdefault(ann["image_id"], []).append(ann)

    with open(MASTER_JSON, "r", encoding="utf-8") as f:
        master = json.load(f)
    master_by_name = {m["book_name"]: m for m in master}

    gt_ref_by_ann: dict[int, str] = {}
    if REF_ASSIGNMENT_CSV.exists():
        with open(REF_ASSIGNMENT_CSV, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                hm = (row.get("hungarian_mask") or "").strip()
                if hm:
                    gt_ref_by_ann[int(hm)] = row["ref"]

    all_rows = []          # 1行 = 1 GTインスタンス
    calib_samples_by_image: dict[int, list[tuple[float, float]]] = {}  # image_id -> [(px, mm), ...]

    # 前回実行分の古いクロップ画像が残らないよう先に掃除する(このスクリプトが新規作成した
    # ディレクトリのみが対象。catheter-100やcatheter/outputs配下の他ファイルは触らない)。
    for old_png in OVERLAY_CROP_DIR.glob("*.png"):
        old_png.unlink()

    # ---- 画像ごとに処理 ----
    for image_id in sorted(images_by_id.keys()):
        if image_id in EXCLUDED_IMAGE_IDS:
            continue
        info = images_by_id[image_id]
        img_path = IMAGES_DIR / info["file_name"]
        height, width = info["height"], info["width"]
        anns = anns_by_image.get(image_id, [])

        cache_path = MASK_CACHE_DIR / f"{image_id}.npz"
        npz = np.load(cache_path)
        pred_masks = [npz[k].astype(np.uint8) for k in sorted(npz.files, key=lambda s: int(s.split("_")[1]))]
        n_pred = len(pred_masks)

        gt_masks = [segmentation_to_mask(a["segmentation"], height, width) for a in anns]
        n_gt = len(anns)

        iou_mat = np.zeros((n_pred, n_gt), dtype=np.float64)
        for i in range(n_pred):
            for j in range(n_gt):
                iou_mat[i, j] = iou(pred_masks[i], gt_masks[j])

        matched_pred_of_gt: dict[int, int] = {}
        if n_pred > 0 and n_gt > 0:
            rows_i, cols_j = linear_sum_assignment(-iou_mat)
            for i, j in zip(rows_i, cols_j):
                if iou_mat[i, j] > 0.0:
                    matched_pred_of_gt[j] = i

        # ---- OCR結果を予測マスクへ帰属させ、マスタと照合してスコア/識別結果を得る ----
        pred_identified: dict[int, str] = {}
        pred_score: dict[int, float] = {}
        pred_text: dict[int, str] = {}
        pred_identified_baseline: dict[int, str] = {}
        pred_score_baseline: dict[int, float] = {}
        ocr_path = OCR_DIR / f"{image_id}.json"
        if ocr_path.exists() and n_pred > 0:
            with open(ocr_path, "r", encoding="utf-8") as f:
                ocr_data = json.load(f)

            def mask_bbox(m):
                ys, xs = np.where(m > 0)
                if ys.size == 0:
                    return None
                return float(xs.min()), float(ys.min()), float(xs.max()), float(ys.max())

            pred_boxes = [mask_bbox(m) for m in pred_masks]

            buckets: list[list[tuple[float, str]]] = [[] for _ in range(n_pred)]
            for item in ocr_data.get("items", []):
                text = (item.get("text") or "").strip()
                if not text or float(item.get("rec_score", 1.0)) < MIN_REC_SCORE:
                    continue
                poly = np.asarray(item["poly_original"], dtype=np.float32)
                xs, ys = poly[:, 0], poly[:, 1]
                mx1, my1, mx2, my2 = xs.min(), ys.min(), xs.max(), ys.max()
                area = (mx2 - mx1) * (my2 - my1)
                if area <= 0:
                    continue
                best_i, best_ratio = None, 0.0
                for i, bb in enumerate(pred_boxes):
                    if bb is None:
                        continue
                    bx1, by1, bx2, by2 = bb
                    ox = max(0.0, min(mx2, bx2) - max(mx1, bx1))
                    oy = max(0.0, min(my2, by2) - max(my1, by1))
                    ratio = (ox * oy) / area
                    if ratio > best_ratio:
                        best_ratio, best_i = ratio, i
                if best_i is not None and best_ratio > 0:
                    buckets[best_i].append((float((my1 + my2) / 2.0), text))

            combined = [" ".join(t for _, t in sorted(b)) for b in buckets]
            for i in range(n_pred):
                pred_text[i] = combined[i]

            n_master = len(master)
            S_ref = np.zeros((n_pred, n_master))
            S_dn = np.zeros((n_pred, n_master))
            S_date = np.zeros((n_pred, n_master))
            for i in range(n_pred):
                c = combined[i]
                for j, mst in enumerate(master):
                    S_ref[i, j] = key_score(mst.get("book_name", ""), c)
                    S_dn[i, j] = key_score(mst.get("display_name", ""), c)
                    S_date[i, j] = key_score(mst.get("expiration date", ""), c, is_date=True)
            S_max = np.maximum(np.maximum(S_ref, S_dn), S_date)

            if n_pred > 0 and n_master > 0:
                rows_i2, cols_j2 = linear_sum_assignment(-S_max)
                for i, j in zip(rows_i2, cols_j2):
                    pred_identified[i] = master[j].get("book_name", "")
                    pred_score[i] = float(S_max[i, j])

                # ---- ベースライン(現行相当): REFキー単独 + マスクごと独立argmax ----
                # multikey_matcher.py(多段query)・match_eval.py(全体最適割当)を使わない場合の再現。
                for i in range(n_pred):
                    j_best = int(np.argmax(S_ref[i]))
                    pred_identified_baseline[i] = master[j_best].get("book_name", "")
                    pred_score_baseline[i] = float(S_ref[i, j_best])

        # ---- 校正: 参照識別(GTマスクベース)が正しいと仮定し、px->mm換算係数を画像ごとに求める ----
        calib_list: list[tuple[float, float]] = []
        for j, ann in enumerate(anns):
            ann_id = ann["id"]
            ref_name = gt_ref_by_ann.get(ann_id, "")
            mst = master_by_name.get(ref_name)
            if mst is None:
                continue
            mm = to_float_or_none(mst.get("book_width"))
            if mm is None:
                continue
            px = rect_width_px(gt_masks[j])
            if px is None or px <= 0:
                continue
            calib_list.append((px, mm))
        calib_samples_by_image[image_id] = calib_list
        if calib_list:
            scale_mm_per_px = float(np.mean([mm / px for px, mm in calib_list]))
        else:
            scale_mm_per_px = None

        # ---- BGR画像読み込み(クロップ画像生成用) ----
        bgr = cv2.imread(str(img_path))

        # ---- インスタンスごとの行を作成 ----
        for j, ann in enumerate(anns):
            ann_id = ann["id"]
            gt_mask = gt_masks[j]
            gt_width_px = rect_width_px(gt_mask)

            i = matched_pred_of_gt.get(j)
            pred_iou = float(iou_mat[i, j]) if i is not None else 0.0
            matched = pred_iou >= IOU_MATCH_THRESHOLD

            identified_name = ""
            display_name = ""
            catalog_width_mm = None
            est_width_mm = None
            width_err_mm = None
            score = None
            ocr_text = ""
            ref_match = ""
            overlay_filename = ""
            identified_name_baseline = ""
            score_baseline = None
            ref_match_baseline = ""

            ref_name = gt_ref_by_ann.get(ann_id, "")

            if matched:
                pred_mask = pred_masks[i]
                pred_width_px = rect_width_px(pred_mask)
                identified_name = pred_identified.get(i, "")
                score = pred_score.get(i)
                ocr_text = pred_text.get(i, "")
                identified_name_baseline = pred_identified_baseline.get(i, "")
                score_baseline = pred_score_baseline.get(i)

                mst = master_by_name.get(identified_name)
                if mst is not None:
                    display_name = mst.get("display_name", "")
                    catalog_width_mm = to_float_or_none(mst.get("book_width"))

                if pred_width_px is not None and scale_mm_per_px is not None:
                    est_width_mm = pred_width_px * scale_mm_per_px
                if est_width_mm is not None and catalog_width_mm is not None:
                    width_err_mm = est_width_mm - catalog_width_mm

                if ref_name and identified_name:
                    ref_match = "一致" if identified_name == ref_name else "不一致"
                if ref_name and identified_name_baseline:
                    ref_match_baseline = "一致" if identified_name_baseline == ref_name else "不一致"

                # ---- クロップ重畳画像 ----
                if bgr is not None:
                    ys, xs = np.where((gt_mask > 0) | (pred_mask > 0))
                    if ys.size > 0:
                        x0 = max(0, int(xs.min()) - CROP_MARGIN)
                        y0 = max(0, int(ys.min()) - CROP_MARGIN)
                        x1 = min(width, int(xs.max()) + CROP_MARGIN)
                        y1 = min(height, int(ys.max()) + CROP_MARGIN)
                        crop = bgr[y0:y1, x0:x1].copy()

                        gt_c, _ = cv2.findContours((gt_mask[y0:y1, x0:x1] > 0).astype(np.uint8),
                                                    cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        cv2.drawContours(crop, gt_c, -1, (0, 255, 0), 2)
                        pr_c, _ = cv2.findContours((pred_mask[y0:y1, x0:x1] > 0).astype(np.uint8),
                                                    cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        cv2.drawContours(crop, pr_c, -1, (0, 0, 255), 2)

                        label = f"#{ann_id:03d} img{image_id} IoU={pred_iou:.2f}"
                        cv2.putText(crop, label, (5, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1, cv2.LINE_AA)

                        overlay_filename = f"{ann_id}-{image_id}-{safe_name(display_name)}.png"
                        cv2.imwrite(str(OVERLAY_CROP_DIR / overlay_filename), crop)

            all_rows.append({
                "認識番号": ann_id,
                "画像番号": image_id,
                "display_name": display_name,
                "book_name": identified_name,
                "book_width_mm": catalog_width_mm,
                "推定幅_mm": round(est_width_mm, 2) if est_width_mm is not None else None,
                "推定幅誤差_mm": round(width_err_mm, 2) if width_err_mm is not None else None,
                "正誤": None,
                "スコア": round(score, 1) if score is not None else None,
                "認識した文字列": ocr_text,
                "IoU": round(pred_iou, 4),
                "参考_GTマスクでの識別結果": ref_name,
                "参考識別との一致": ref_match,
                "book_name_ベースライン(現行相当)": identified_name_baseline,
                "スコア_ベースライン": round(score_baseline, 1) if score_baseline is not None else None,
                "参考識別との一致_ベースライン": ref_match_baseline,
                "オーバーレイ画像": overlay_filename,
            })

    # ---- Excel 出力 ----
    wb = Workbook()
    ws = wb.active
    ws.title = "recognition_report"

    headers = list(all_rows[0].keys())
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(horizontal="center")

    for row in all_rows:
        ws.append([row[h] for h in headers])

    last_row = len(all_rows) + 1
    last_col = len(headers)

    # オートフィルタ(全列で並び替え/絞り込み可能に)
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"

    # 正誤列に T/F のドロップダウン(データ入力規則)を設定
    seiigo_col = headers.index("正誤") + 1
    seiigo_letter = get_column_letter(seiigo_col)
    dv = DataValidation(type="list", formula1='"T,F"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{seiigo_letter}2:{seiigo_letter}{last_row}")

    # 列幅
    widths = {
        "認識番号": 10, "画像番号": 10, "display_name": 26, "book_name": 18,
        "book_width_mm": 13, "推定幅_mm": 12, "推定幅誤差_mm": 14, "正誤": 8,
        "スコア": 10, "認識した文字列": 50, "IoU": 10,
        "参考_GTマスクでの識別結果": 20, "参考識別との一致": 14,
        "book_name_ベースライン(現行相当)": 24, "スコア_ベースライン": 14,
        "参考識別との一致_ベースライン": 18, "オーバーレイ画像": 20,
    }
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 14)

    out_xlsx = EVAL_DIR / "catheter_recognition_report.xlsx"
    wb.save(out_xlsx)
    print(f"✔ Excel report -> {out_xlsx}  ({len(all_rows)} rows)")
    print(f"✔ overlay crops -> {OVERLAY_CROP_DIR}  ({len(list(OVERLAY_CROP_DIR.glob('*.png')))} files)")

    # ---- ベースライン(現行相当: REF単独+独立argmax) vs 改善後(多段query+ハンガリー法) 比較 ----
    comparable = [r for r in all_rows
                  if r["参考_GTマスクでの識別結果"] and r["book_name"] and r["book_name_ベースライン(現行相当)"]]
    if comparable:
        n_cmp = len(comparable)
        n_ok_improved = sum(1 for r in comparable if r["参考識別との一致"] == "一致")
        n_ok_baseline = sum(1 for r in comparable if r["参考識別との一致_ベースライン"] == "一致")
        lines = [
            "# 識別ロジックのアブレーション比較 (multikey_matcher.py / match_eval.py の効果)",
            "",
            f"比較可能件数: {n_cmp}",
            f"- ベースライン(現行相当: REFキー単独 + マスクごと独立argmax): "
            f"{n_ok_baseline} / {n_cmp} ({100.0*n_ok_baseline/n_cmp:.1f}%)",
            f"- 改善後(REF+display_name+期限の多段query + 全体最適割当ハンガリー法): "
            f"{n_ok_improved} / {n_cmp} ({100.0*n_ok_improved/n_cmp:.1f}%)",
            f"- 差分: {n_ok_improved - n_ok_baseline:+d}件 "
            f"({100.0*(n_ok_improved-n_ok_baseline)/n_cmp:+.1f}pt)",
            "",
            "※ IoU・幅(px/mm)はどちらの識別ロジックでも変化しない(SAMのマスク自体は同じため)。"
            "変わるのは識別結果(正誤)のみ。",
        ]
        ablation_path = EVAL_DIR / "identification_ablation.md"
        ablation_path.write_text("\n".join(lines), encoding="utf-8")
        print("\n" + "\n".join(lines))
        print(f"\n✔ ablation summary -> {ablation_path}")


if __name__ == "__main__":
    main()
