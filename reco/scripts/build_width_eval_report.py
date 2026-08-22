#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
width_mm_validation.py の実行結果(reco/<dataset>/width_eval_result.csv)をもとに、
データセットごとに目視レビュー用のExcelと、各件のfinal.png(選択された箱をハイライトした
画像)を集めた画像フォルダを作る。

出力先(データセットごとに分離、実行日時ごとに新規フォルダ。2026-08-21、ユーザー要望:
Excel/画像が毎回上書きされず区別できるようにするため):
  reco/<dataset>/reco_result_<実行日時>/images/*.png
  reco/<dataset>/reco_result_<実行日時>/catheter_width_report_<実行日時>.xlsx

Excelと画像フォルダは同じreco_result_<実行日時>/の直下にまとめる(2026-08-21、
ユーザー要望: 「Excelと画像が別フォルダに分かれているのはNG、1つのフォルダに
まとめてほしい」)。認識処理自体の作業フォルダwidth_eval_work<suffix>/はこの
レポート生成の出力先とは別物で、入力データの読み取り元としてのみ使う(変更なし)。

reco_result_<実行日時>/work は width_eval_work<suffix>/ へのシンボリックリンク
(2026-08-21、ユーザー要望: 「各アイテムの個別認識情報(final.png以外にocr_result.json・
debug_*等)にもreco_result側からアクセスしたい」)。width_eval_work<suffix>/は
stand-100だけで4GB超あり、レポート生成のたびに実体コピーすると容量が急増するため、
シンボリックリンクで参照する形にした(実体コピーではない)。

実行:
    cd ~/pro_book/pro_hand_book_python
    .pro_hand_book_fixed/bin/python3.10 reco/scripts/build_width_eval_report.py
    .pro_hand_book_fixed/bin/python3.10 reco/scripts/build_width_eval_report.py --suffix _rotfix
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import time
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

RECO_ROOT = Path(__file__).resolve().parents[1]
DATASETS = ["stand-100", "diagonal-40"]

# データセットごとのCVAT/COCO形式アノテーション(GTポリゴン、IoU一致度の算出に使う)。
ANNOTATIONS_JSON = {
    "stand-100": RECO_ROOT / "stand-100" / "annotations" / "instances_default_100.json",
    "diagonal-40": RECO_ROOT / "diagonal-40" / "annotations" / "instances_default.json",
}

# パイプラインごとに最終選択マスクの保存ファイル名が異なる(2026-08-22判明:
# simplifiedパイプラインはfinal_mask.pngを作らずselected_mask_refined.pngを使う)。
# 存在する方を優先順に試す。
SELECTED_MASK_FILENAMES = ["selected_mask_refined.png", "final_mask.png", "selected_mask_raw.png"]

HEADERS = [
    "認識した順番", "画像ファイル", "shot", "query(book_name)", "display_name",
    "目視確認(T/F)", "IoU一致度",
    "正解幅mm", "推定幅mm", "誤差mm", "2mm以内(把持成功目安)", "リトライ回数",
    "識別スコア", "識別margin", "識別確信度(confident)", "認識した文字列",
    "処理時間sec", "メモ", "エラー",
]
COLUMN_WIDTHS = {
    "認識した順番": 12, "画像ファイル": 42, "shot": 20, "query(book_name)": 16,
    "display_name": 26, "目視確認(T/F)": 14, "IoU一致度": 12,
    "正解幅mm": 10, "推定幅mm": 10, "誤差mm": 10,
    "2mm以内(把持成功目安)": 18, "リトライ回数": 12, "識別スコア": 12, "識別margin": 12,
    "識別確信度(confident)": 16, "認識した文字列": 40, "処理時間sec": 12,
    "メモ": 24, "エラー": 20,
}


def safe_name(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in s)


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def image_filename_for(dataset: str, shot: str, display_name: str, work_dir_name: str = "") -> str:
    """データセットごとの分かりやすい画像ファイル名を作る。

    diagonal-40のshot名は元々品目名そのもの(例: Target_R)で読みやすいが、
    stand-100のshot名は"<画像番号>__<book_nameコード>"(例: 1__ESC0305)で
    コードが読みにくいため、display_nameベースの名前に置き換える。

    stand-100・diagonal-40とも、実際に解決された作業フォルダ名(work_dir_name、例:
    `1-1-Target_XL`・`23-AXS_DAC_L`)をそのまま使う(2026-08-21/22、ユーザー要望:
    「images内の画像の命名規則もフォルダと同様にしてください」)。こうすることで
    区切り文字も含めてworkフォルダ名と常に一致することが保証される(別々に
    safe_name組み立てをやり直すと表記がずれる恐れがあるため、単一のソース=
    work_dir.nameから作る)。
    """
    if work_dir_name:
        return f"{work_dir_name}.png"
    if dataset == "stand-100":
        return f"{safe_name(display_name or shot)}.png"
    return f"{safe_name(shot)}.png"


def resolve_work_dir(base_dir: Path, dataset: str, shot: str, display_name: str = "") -> Path:
    """shotに対応する作業フォルダを解決する。

    stand-100のwidth_eval_work_rotfixは、shot名(`{画像番号}__{REF}`)のまま
    ではなく`{認識番号}-{画像番号}-{display_name}`形式にリネーム済み(2026-08-21、
    別タスクで実施)なので、直接一致しない場合はこのパターンでフォールバック探索する。
    diagonal-40も2026-08-22から作業フォルダ名の先頭に処理順を付けるようになった
    (`{処理順}-{shot名}`)ため、同様にフォールバック探索する。

    【2026-08-21 重大バグ修正】画像番号のみ(例: `^\\d+-1-`)で前方一致させていたため、
    同じ画像番号を共有する20件全部が同じフォルダ(ソート順で最初に見つかったもの)に
    解決されてしまい、Excelの識別スコア以降の列が20件ずつ同じ値になっていた
    (ユーザー報告により発覚)。display_nameのsafe_name化した値まで完全一致させる
    ことで、1件ずつ正しく一意なフォルダに解決するよう修正した。
    """
    direct = base_dir / shot
    if direct.exists():
        return direct
    if dataset == "stand-100" and "__" in shot and display_name:
        image_id = shot.split("__", 1)[0]
        expected = f"{image_id}-{safe_name(display_name)}"
        pattern = re.compile(rf"^\d+-{re.escape(expected)}$")
        matches = [
            cand for cand in (sorted(base_dir.iterdir()) if base_dir.exists() else [])
            if cand.is_dir() and pattern.match(cand.name)
        ]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            print(f"⚠ resolve_work_dir: shot={shot} display_name={display_name!r} に対して"
                  f"複数候補が一致し一意に決まりません: {[m.name for m in matches]}")
    elif dataset == "diagonal-40":
        expected = safe_name(shot)
        pattern = re.compile(rf"^\d+-{re.escape(expected)}$")
        matches = [
            cand for cand in (sorted(base_dir.iterdir()) if base_dir.exists() else [])
            if cand.is_dir() and pattern.match(cand.name)
        ]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            print(f"⚠ resolve_work_dir: shot={shot} に対して"
                  f"複数候補が一致し一意に決まりません: {[m.name for m in matches]}")
    return direct


def load_multikey_debug(work_dir: Path) -> dict:
    p = work_dir / "multikey_match_debug.json"
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def recognized_text_for_selected_mask(debug: dict) -> str:
    selected_mask = debug.get("selected_mask")
    for m in debug.get("per_mask", []):
        if m.get("mask") == selected_mask:
            return m.get("text", "")
    return ""


def decode_uncompressed_rle(counts: list[int], height: int, width: int) -> np.ndarray:
    """COCOの非圧縮RLE(countsがintのリスト)を0/255マスクにデコードする。列優先(Fortran順)で
    0/1が交互に続くランレングス形式(reco/stand-100/scripts/compare_quad_fit.pyの同名関数と同じ)。"""
    flat = np.zeros(height * width, dtype=np.uint8)
    idx = 0
    val = 0
    for c in counts:
        if c:
            flat[idx: idx + c] = val * 255
        idx += c
        val = 1 - val
    return flat.reshape((height, width), order="F")


def segmentation_to_mask(segmentation, height: int, width: int) -> np.ndarray:
    """COCOセグメンテーション(ポリゴン形式 or 非圧縮RLE形式)を0/255マスクにラスタライズする。"""
    if isinstance(segmentation, dict):
        counts = segmentation["counts"]
        if isinstance(counts, str):
            raise NotImplementedError(
                "圧縮RLE(countsが文字列)は未対応です(pycocotools未導入のため)。"
            )
        return decode_uncompressed_rle(counts, height, width)

    mask = np.zeros((height, width), dtype=np.uint8)
    for part in segmentation:
        pts = np.array(part, dtype=np.float64).reshape(-1, 2)
        pts = np.round(pts).astype(np.int32).reshape(-1, 1, 2)
        cv2.fillPoly(mask, [pts], 255)
    return mask


def load_gt_masks_by_image(dataset: str) -> dict[str, list[np.ndarray]]:
    """アノテーションJSON(CVAT/COCO形式)を読み、画像を特定するキー->GTマスク一覧、を返す。

    stand-100は画像ファイル名が"1.png"〜"5.png"なのでキーは画像番号("1"等)、
    diagonal-40は画像ファイル名がshot名そのもの("AXS_DAC_L.png"等)なのでキーはshot名。
    1画像につき複数(stand-100は20個/枚)のGTポリゴンがあるので、IoU一致度は
    「選択マスクと、その画像内の全GTポリゴンとの最大IoU」とする(2026-08-22、
    ユーザー要望: 「目視確認の右隣にIoU一致度を入れたい」。個々のGTがどのbook_nameに
    対応するかの正解ラベル付けは無いため、識別の正誤とは独立に「セグメンテーション
    そのものの幾何精度」を見る指標として算出する)。
    """
    ann_path = ANNOTATIONS_JSON.get(dataset)
    if not ann_path or not ann_path.exists():
        return {}
    try:
        d = json.loads(ann_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    images_by_id = {img["id"]: img for img in d.get("images", [])}
    anns_by_image_id: dict[int, list] = {}
    for ann in d.get("annotations", []):
        anns_by_image_id.setdefault(ann["image_id"], []).append(ann)

    result: dict[str, list[np.ndarray]] = {}
    for image_id, img in images_by_id.items():
        file_name = img.get("file_name", "")
        key = Path(file_name).stem  # "1.png"->"1", "AXS_DAC_L.png"->"AXS_DAC_L"
        h, w = img.get("height"), img.get("width")
        if not h or not w:
            continue
        masks = []
        for ann in anns_by_image_id.get(image_id, []):
            seg = ann.get("segmentation")
            if not seg:
                continue
            try:
                mask = segmentation_to_mask(seg, h, w)
                # annotations/(images/*.png)はdepth_shots側と180度向きが異なる
                # (2026-08-21、ユーザー確認済み: depth_shotsが本来正しい向きで、
                # annotations側の方を回転させる必要がある。実測でも無回転だと
                # 最良IoUが0.357、180度回転後は0.882まで跳ね上がることを確認済み)。
                mask = np.rot90(mask, 2)
                masks.append(mask)
            except NotImplementedError as e:
                print(f"⚠ GTアノテーション(id={ann.get('id')})のIoU算出をスキップ: {e}")
        result[key] = masks
    return result


def gt_image_key_for_shot(dataset: str, shot: str) -> str:
    if dataset == "stand-100" and "__" in shot:
        return shot.split("__", 1)[0]
    return shot


def load_selected_mask(work_dir: Path) -> np.ndarray | None:
    for name in SELECTED_MASK_FILENAMES:
        p = work_dir / name
        if p.exists():
            m = cv2.imread(str(p), cv2.IMREAD_UNCHANGED)
            if m is not None:
                return m
    return None


def compute_iou(mask_a: np.ndarray, mask_b: np.ndarray) -> float | None:
    a = mask_a > 0
    b = mask_b > 0
    if a.shape != b.shape:
        return None
    inter = np.logical_and(a, b).sum()
    union = np.logical_or(a, b).sum()
    if union == 0:
        return None
    return float(inter) / float(union)


def best_iou_for_row(dataset: str, shot: str, work_dir: Path,
                      gt_masks_by_image: dict[str, list[np.ndarray]]) -> float | None:
    gt_masks = gt_masks_by_image.get(gt_image_key_for_shot(dataset, shot))
    if not gt_masks:
        return None
    selected = load_selected_mask(work_dir)
    if selected is None:
        return None
    ious = [iou for gt in gt_masks if (iou := compute_iou(selected, gt)) is not None]
    return max(ious) if ious else None


def build_dataset_report(dataset: str, suffix: str = "", report_timestamp: str = "") -> None:
    """suffixを指定すると、入力データは width_eval_result{suffix}.csv /
    width_eval_work{suffix}/ から読む(2026-08-21、回転バグ修正版=_rotfixを反映する際に追加)。

    xlsxと画像は両方とも reco_result_{report_timestamp}/ の直下にまとまって入る
    (xlsx本体とimages/サブフォルダ)。実行のたびに新しいフォルダ・ファイル名になり、
    Excel・画像とも上書きされない(2026-08-21、ユーザー要望:
    「Excelファイルが毎回上書きされるのは困る」「画像が入るフォルダ名の区別もつきにくい」
    への対応)。認識処理の作業フォルダwidth_eval_work{suffix}/は入力データの読み取り元
    としてのみ使う。reco_result_{report_timestamp}/work はこのwidth_eval_work{suffix}/への
    シンボリックリンク(2026-08-21、ユーザー要望: 「個別アイテムの詳細情報にも
    reco_result側からアクセスしたい」。実体コピーだと容量が急増するためリンクにした)。
    """
    csv_path = RECO_ROOT / dataset / f"width_eval_result{suffix}.csv"
    if not csv_path.exists():
        print(f"⚠ 見つかりません、スキップ: {csv_path}")
        return

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    # 認識した順番 = CSVへの追記順(=実際に処理された順、resume実行をまたいでも保たれる)。
    # stand-100のwork/フォルダ名・images/ファイル名の先頭数字と同じ値(2026-08-21、
    # ユーザー要望: 「エクセルの一番左列に認識した順番を加えてほしい」)。
    for i, r in enumerate(rows):
        r["_proc_order"] = i + 1
    rows.sort(key=lambda r: r["shot"])

    out_dir = RECO_ROOT / dataset / f"reco_result_{report_timestamp}"
    work_base_dir = RECO_ROOT / dataset / f"width_eval_work{suffix}"
    images_dir = out_dir / "images"
    xlsx_path = out_dir / f"catheter_width_report_{report_timestamp}.xlsx"
    images_dir.mkdir(parents=True, exist_ok=True)

    work_link = out_dir / "work"
    if work_base_dir.exists() and not work_link.exists():
        work_link.symlink_to(
            Path(os.path.relpath(work_base_dir, out_dir)), target_is_directory=True
        )

    gt_masks_by_image = load_gt_masks_by_image(dataset)

    for r in rows:
        shot = r["shot"]
        work_dir = resolve_work_dir(work_base_dir, dataset, shot, r.get("display_name", ""))
        debug = load_multikey_debug(work_dir)
        r["_selected_score"] = debug.get("selected_score")
        r["_margin"] = debug.get("margin")
        r["_confident"] = debug.get("confident")
        r["_recognized_text"] = recognized_text_for_selected_mask(debug)
        r["_iou"] = best_iou_for_row(dataset, shot, work_dir, gt_masks_by_image)

        img_filename = image_filename_for(dataset, shot, r.get("display_name", ""), work_dir.name)
        src = work_dir / "final.png"
        r["_image_filename"] = ""
        if src.exists():
            shutil.copyfile(src, images_dir / img_filename)
            r["_image_filename"] = img_filename
        else:
            print(f"⚠ final.pngが見つかりません: {src}")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_name(dataset)[:31]

    ws.append(HEADERS)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for r in rows:
        true_mm = to_float(r.get("book_width_mm_true"))
        pred_mm = to_float(r.get("book_width_mm_pred"))
        err_mm = to_float(r.get("abs_error_mm"))
        ws.append([
            r["_proc_order"],
            r["_image_filename"],
            r["shot"],
            r.get("query", ""),
            r.get("display_name", ""),
            "",
            (round(r["_iou"], 3) if r["_iou"] is not None else ""),
            true_mm,
            pred_mm,
            err_mm,
            ("○" if err_mm is not None and err_mm <= 2.0 else ("" if err_mm is None else "×")),
            to_float(r.get("retry_count")),
            to_float(r["_selected_score"]),
            to_float(r["_margin"]),
            ("" if r["_confident"] is None else str(r["_confident"])),
            r["_recognized_text"],
            to_float(r.get("elapsed_sec")),
            "",
            r.get("error", ""),
        ])

    n_rows = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{n_rows}"
    for idx, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(h, 14)

    review_col = HEADERS.index("目視確認(T/F)") + 1
    review_letter = get_column_letter(review_col)
    dv = DataValidation(type="list", formula1='"T,F"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{review_letter}2:{review_letter}{n_rows}")

    err_col = HEADERS.index("誤差mm") + 1
    for row_idx in range(2, n_rows + 1):
        err_cell = ws.cell(row=row_idx, column=err_col)
        if isinstance(err_cell.value, (int, float)):
            if err_cell.value <= 2.0:
                err_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif err_cell.value >= 10.0:
                err_cell.fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")

    wb.save(xlsx_path)
    print(f"✔ [{dataset}] Excel -> {xlsx_path} ({len(rows)}行)")
    print(f"✔ [{dataset}] 画像 -> {images_dir} ({len(list(images_dir.glob('*.png')))}枚)")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--suffix", default="",
                    help="入力データの接尾辞(例: _rotfix、width_mm_validation.pyの"
                         "--work-suffixに対応)。")
    args = ap.parse_args()
    report_timestamp = time.strftime("%Y%m%d_%H%M%S")
    for dataset in DATASETS:
        build_dataset_report(dataset, suffix=args.suffix, report_timestamp=report_timestamp)


if __name__ == "__main__":
    main()
