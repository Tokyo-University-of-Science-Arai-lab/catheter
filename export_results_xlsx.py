#!/usr/bin/env python3
"""
offline_pointcloud_debug_SAM3.py の results.csv を Excel (.xlsx) に変換する。

出力: <run_dir>/results.xlsx
  ・display_name（書籍名）の列を追加
  ・数値は数値として書き込む（Excel上で並べ替え・集計できる）
  ・見出し固定、オートフィルタ、列幅調整あり
  ・誤差の大きい行に色を付ける（軸反転の発見用）

使い方:
  python export_results_xlsx.py                # 最新の実行結果を変換
  python export_results_xlsx.py <run_dir>      # 実行結果を指定
"""

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
OFFLINE_BASE_DIR = BASE_DIR / "captures" / "5shot_catheter_offline"
MASTER_JSON = BASE_DIR / "master_20260216.json"

# 誤差がこれを超えたら「軸反転の疑い」として色を付ける [mm]
FLIP_ERROR_MM = 50.0
# 誤差がこれ以内なら「良好」として色を付ける [mm]
GOOD_ERROR_MM = 2.0

# (CSVの列名, Excelの見出し, 数値として扱うか, 列幅)
COLUMNS = [
    ("trial_no",            "試行",           True,   6),
    ("shot_id",             "画像",           True,   6),
    ("master_index",        "品目番号",       True,   9),
    ("display_name",        "書籍名",         False, 32),
    ("book_name",           "型番",           False, 18),
    ("gt_book_width_mm",    "正解幅[mm]",     True,  11),
    ("pred_book_width_mm",  "推定幅[mm]",     True,  11),
    ("signed_error_mm",     "誤差[mm]",       True,  11),
    ("abs_error_mm",        "絶対誤差[mm]",   True,  12),
    ("roll_rad",            "roll[rad]",      True,  11),
    ("elapsed_sec",         "所要[秒]",       True,   9),
    ("status",              "状態",           False,  9),
    ("error",               "エラー内容",     False, 30),
    ("run_shot_dir",        "出力フォルダ",   False, 60),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
FLIP_FILL   = PatternFill("solid", fgColor="FFC7CE")   # 赤系: 軸反転の疑い
GOOD_FILL   = PatternFill("solid", fgColor="C6EFCE")   # 緑系: 良好
FAIL_FILL   = PatternFill("solid", fgColor="FFEB9C")   # 黄系: 認識失敗


def latest_run_dir(base_dir: Path) -> Path:
    runs = [p for p in base_dir.iterdir() if p.is_dir() and (p / "results.csv").exists()]
    if not runs:
        raise FileNotFoundError(f"results.csv を持つ実行結果がありません: {base_dir}")
    return max(runs, key=lambda p: p.stat().st_mtime)


def load_display_names(master_json: Path) -> dict[str, str]:
    data = json.loads(master_json.read_text(encoding="utf-8"))
    return {
        item["book_name"]: item.get("display_name") or item["book_name"]
        for item in data
    }


def to_number(value):
    """数値に変換できればfloat/intで返す。できなければ元の文字列のまま返す。"""
    if value in (None, ""):
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return value
    return int(f) if f.is_integer() and abs(f) < 1e9 else f


def main():
    run_dir = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else latest_run_dir(OFFLINE_BASE_DIR)
    results_csv = run_dir / "results.csv"
    if not results_csv.exists():
        raise FileNotFoundError(f"results.csv がありません: {results_csv}")

    display_names = load_display_names(MASTER_JSON)

    with open(results_csv, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # 見出し
    ws.append([label for _, label, _, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 本体
    for r in rows:
        r = dict(r)
        r["display_name"] = display_names.get(r.get("book_name", ""), r.get("book_name", ""))
        ws.append([
            to_number(r.get(key)) if numeric else r.get(key, "")
            for key, _, numeric, _ in COLUMNS
        ])

    # 列幅・数値の表示形式
    error_cols = {"signed_error_mm", "abs_error_mm", "pred_book_width_mm",
                  "gt_book_width_mm", "roll_rad", "elapsed_sec"}
    for i, (key, _, _, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = width
        if key in error_cols:
            for cell in ws[letter][1:]:
                cell.number_format = "0.000"

    # 誤差に応じて行を色分けする
    abs_col = next(i for i, (k, _, _, _) in enumerate(COLUMNS, start=1) if k == "abs_error_mm")
    n_flip = n_good = n_fail = 0
    for row in ws.iter_rows(min_row=2):
        status = row[next(i for i, (k, _, _, _) in enumerate(COLUMNS) if k == "status")].value
        err = row[abs_col - 1].value

        if status != "success":
            fill, hit = FAIL_FILL, "fail"
        elif isinstance(err, (int, float)) and err > FLIP_ERROR_MM:
            fill, hit = FLIP_FILL, "flip"
        elif isinstance(err, (int, float)) and err <= GOOD_ERROR_MM:
            fill, hit = GOOD_FILL, "good"
        else:
            continue

        for cell in row:
            cell.fill = fill
        n_fail += hit == "fail"
        n_flip += hit == "flip"
        n_good += hit == "good"

    # 見出し固定とオートフィルタ
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path = run_dir / "results.xlsx"
    wb.save(out_path)

    print(f"対象: {results_csv}")
    print(f"出力: {out_path}")
    print(f"  {len(rows)} 行")
    print(f"  緑（誤差{GOOD_ERROR_MM}mm以内）: {n_good}")
    print(f"  赤（誤差{FLIP_ERROR_MM}mm超・軸反転の疑い）: {n_flip}")
    print(f"  黄（認識失敗）: {n_fail}")


if __name__ == "__main__":
    main()
