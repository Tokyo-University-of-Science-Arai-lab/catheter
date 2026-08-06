#!/usr/bin/env python3
"""
offline_pointcloud_debug_SAM3.py の結果から、精度をまとめた Excel を作る。

results.xlsx（export_results_xlsx.py）が全試行の一覧なのに対し、
こちらは集計を主にする。シートは4つ。

  概要    しきい値ごとの割合・誤差の統計・実行時間
  品目別  20品目それぞれの成績
  画像別  shotごとの成績
  全試行  100試行の一覧（誤差で色分け）

比較対象の実行を渡すと、概要・品目別に「前→後」の列が入る。

使い方:
  python export_accuracy_xlsx.py                          # 最新の実行結果
  python export_accuracy_xlsx.py <run_dir>                # 実行結果を指定
  python export_accuracy_xlsx.py <run_dir> <compare_dir>  # 比較対象も指定
"""

import json
import statistics
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
OFFLINE_BASE_DIR = BASE_DIR / "captures" / "5shot_catheter_offline"
MASTER_JSON = BASE_DIR / "master_20260216.json"

# 誤差がこれを超えたら「軸反転の疑い」として扱う [mm]
FLIP_ERROR_MM = 50.0
# 集計するしきい値 [mm]
THRESHOLDS_MM = [1.0, 1.5, 2.0, 3.0, 5.0, 10.0]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FILL = PatternFill("solid", fgColor="F2F2F2")
FLIP_FILL = PatternFill("solid", fgColor="FFC7CE")   # 赤系: 軸反転の疑い
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")   # 緑系: 2mm以内
FAIL_FILL = PatternFill("solid", fgColor="FFEB9C")   # 黄系: 認識失敗

BOLD = Font(bold=True)


def latest_run_dir(base_dir: Path) -> Path:
    runs = [p for p in base_dir.iterdir() if p.is_dir() and (p / "results.json").exists()]
    if not runs:
        raise FileNotFoundError(f"results.json を持つ実行結果がありません: {base_dir}")
    return max(runs, key=lambda p: p.stat().st_mtime)


def load_display_names(master_json: Path) -> dict[str, str]:
    data = json.loads(master_json.read_text(encoding="utf-8"))
    return {
        item["book_name"]: item.get("display_name") or item["book_name"]
        for item in data
    }


def load_run(run_dir: Path):
    """results.json と、あれば前処理の時間を読む。"""
    results = json.loads((run_dir / "results.json").read_text(encoding="utf-8"))
    prepare_path = run_dir / "_prepared" / "prepare_timing.json"
    prepare = (
        json.loads(prepare_path.read_text(encoding="utf-8"))
        if prepare_path.exists() else []
    )
    return results, prepare


def errors_of(results):
    return [r["abs_error_mm"] for r in results if r["abs_error_mm"] is not None]


def within(results, threshold_mm: float) -> int:
    return sum(1 for e in errors_of(results) if e < threshold_mm)


def write_header(ws, labels, row=1):
    ws.append(labels)
    for cell in ws[row]:
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_overview(wb, run_dir, results, prepare, compare, compare_dir):
    ws = wb.create_sheet("概要")
    n = len(results)

    def title(text):
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = BOLD
        cell.fill = TITLE_FILL

    title("実行情報")
    ws.append(["実行ディレクトリ", str(run_dir)])
    ws.append(["試行数", n])
    ws.append(["完走", sum(1 for r in results if r["status"] == "success")])
    if compare is not None:
        ws.append(["比較対象", str(compare_dir)])
    ws.append([])

    title("精度")
    header = ["しきい値", "件数", "割合"]
    if compare is not None:
        header += ["比較対象の件数", "比較対象の割合", "増減"]
    write_header(ws, header, row=ws.max_row + 1)
    for th in THRESHOLDS_MM:
        c = within(results, th)
        row = [f"{th:.1f}mm未満", c, c / n if n else None]
        if compare is not None:
            co = within(compare, th)
            row += [co, co / len(compare) if compare else None, c - co]
        ws.append(row)
    ws.append([])

    title("誤差の統計")
    e = errors_of(results)
    eo = errors_of(compare) if compare is not None else None
    stats = [
        ("平均絶対誤差[mm]", statistics.mean(e) if e else None),
        ("中央値絶対誤差[mm]", statistics.median(e) if e else None),
        ("最小絶対誤差[mm]", min(e) if e else None),
        ("最大絶対誤差[mm]", max(e) if e else None),
        ("誤差50mm超（軸反転の疑い）[件]", sum(1 for x in e if x > FLIP_ERROR_MM)),
        ("過小推定[件]", sum(1 for r in results
                          if r["signed_error_mm"] is not None and r["signed_error_mm"] < 0)),
        ("過大推定[件]", sum(1 for r in results
                          if r["signed_error_mm"] is not None and r["signed_error_mm"] > 0)),
    ]
    stats_old = None
    if eo is not None:
        stats_old = [
            statistics.mean(eo) if eo else None,
            statistics.median(eo) if eo else None,
            min(eo) if eo else None,
            max(eo) if eo else None,
            sum(1 for x in eo if x > FLIP_ERROR_MM),
            sum(1 for r in compare
                if r["signed_error_mm"] is not None and r["signed_error_mm"] < 0),
            sum(1 for r in compare
                if r["signed_error_mm"] is not None and r["signed_error_mm"] > 0),
        ]
    write_header(ws, ["項目", "値"] + (["比較対象"] if compare is not None else []),
                 row=ws.max_row + 1)
    for i, (label, value) in enumerate(stats):
        row = [label, value]
        if stats_old is not None:
            row.append(stats_old[i])
        ws.append(row)
    ws.append([])

    title("実行時間")
    elapsed = [r["elapsed_sec"] for r in results if r.get("elapsed_sec")]
    prep = [p["prepare_seconds"] for p in prepare]
    ws.append(["1試行あたり[秒]", statistics.mean(elapsed) if elapsed else None])
    if prep:
        ws.append(["前処理（OCR+SAM3）1回あたり[秒]", statistics.mean(prep)])
        ws.append(["前処理の回数", len(prep)])
        ws.append(["画像1枚あたり[秒]",
                   statistics.mean(prep) + statistics.mean(elapsed) * (n / max(len(prep), 1))])
    ws.append(["総計[分]", (sum(elapsed) + sum(prep)) / 60])

    set_widths(ws, [34, 18, 16, 18, 18, 12])
    for row in ws.iter_rows(min_col=3, max_col=5):
        for cell in row:
            if isinstance(cell.value, float) and 0.0 <= cell.value <= 1.0:
                cell.number_format = "0.0%"
    for row in ws.iter_rows(min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, float) and cell.number_format == "General":
                cell.number_format = "0.000"
    return ws


def sheet_per_book(wb, results, compare, display_names):
    ws = wb.create_sheet("品目別")
    header = ["品目番号", "型番", "書籍名", "正解幅[mm]", "試行数",
              "2mm未満", "平均誤差[mm]", "中央値[mm]", "最大誤差[mm]", "最小誤差[mm]",
              "軸反転の疑い"]
    if compare is not None:
        header += ["比較 2mm未満", "比較 平均誤差[mm]", "比較 軸反転"]
    write_header(ws, header)

    by = {}
    for r in results:
        by.setdefault((r["master_index"], r["book_name"], r["gt_book_width_mm"]), []).append(r)
    by_old = {}
    if compare is not None:
        for r in compare:
            by_old.setdefault(r["master_index"], []).append(r)

    for (idx, name, gt), v in sorted(by.items()):
        e = errors_of(v)
        row = [
            idx, name, display_names.get(name, name), gt, len(v),
            sum(1 for x in e if x < 2.0),
            statistics.mean(e) if e else None,
            statistics.median(e) if e else None,
            max(e) if e else None,
            min(e) if e else None,
            sum(1 for x in e if x > FLIP_ERROR_MM),
        ]
        if compare is not None:
            eo = errors_of(by_old.get(idx, []))
            row += [
                sum(1 for x in eo if x < 2.0),
                statistics.mean(eo) if eo else None,
                sum(1 for x in eo if x > FLIP_ERROR_MM),
            ]
        ws.append(row)

    set_widths(ws, [9, 18, 32, 11, 8, 9, 13, 12, 13, 13, 12, 12, 16, 10])
    for col in ("D", "G", "H", "I", "J", "M"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def sheet_per_shot(wb, results):
    ws = wb.create_sheet("画像別")
    write_header(ws, ["画像", "試行数", "2mm未満", "割合",
                      "平均誤差[mm]", "中央値[mm]", "最大誤差[mm]", "軸反転の疑い"])
    by = {}
    for r in results:
        by.setdefault(r["shot_id"], []).append(r)
    for shot_id, v in sorted(by.items()):
        e = errors_of(v)
        ok = sum(1 for x in e if x < 2.0)
        ws.append([
            shot_id, len(v), ok, ok / len(v) if v else None,
            statistics.mean(e) if e else None,
            statistics.median(e) if e else None,
            max(e) if e else None,
            sum(1 for x in e if x > FLIP_ERROR_MM),
        ])
    set_widths(ws, [8, 8, 10, 8, 13, 12, 13, 12])
    for col in ("E", "F", "G"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    for cell in ws["D"][1:]:
        cell.number_format = "0.0%"
    return ws


def sheet_all_trials(wb, results, display_names):
    ws = wb.create_sheet("全試行")
    write_header(ws, ["試行", "画像", "品目番号", "書籍名", "型番",
                      "正解幅[mm]", "推定幅[mm]", "誤差[mm]", "絶対誤差[mm]",
                      "roll[rad]", "所要[秒]", "状態"])
    for r in results:
        ws.append([
            r["trial_no"], r["shot_id"], r["master_index"],
            display_names.get(r["book_name"], r["book_name"]), r["book_name"],
            r["gt_book_width_mm"], r["pred_book_width_mm"],
            r["signed_error_mm"], r["abs_error_mm"], r["roll_rad"],
            r["elapsed_sec"], r["status"],
        ])

    for row in ws.iter_rows(min_row=2):
        status = row[11].value
        err = row[8].value
        if status != "success":
            fill = FAIL_FILL
        elif isinstance(err, (int, float)) and err > FLIP_ERROR_MM:
            fill = FLIP_FILL
        elif isinstance(err, (int, float)) and err <= 2.0:
            fill = GOOD_FILL
        else:
            continue
        for cell in row:
            cell.fill = fill

    set_widths(ws, [6, 6, 9, 32, 18, 11, 11, 11, 12, 11, 9, 9])
    for col in ("F", "G", "H", "I", "J", "K"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    run_dir = Path(args[0]).resolve() if args else latest_run_dir(OFFLINE_BASE_DIR)
    compare_dir = Path(args[1]).resolve() if len(args) > 1 else None

    results, prepare = load_run(run_dir)
    compare = load_run(compare_dir)[0] if compare_dir else None
    display_names = load_display_names(MASTER_JSON)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb, run_dir, results, prepare, compare, compare_dir)
    sheet_per_book(wb, results, compare, display_names)
    sheet_per_shot(wb, results)
    sheet_all_trials(wb, results, display_names)

    out_path = run_dir / "accuracy_summary.xlsx"
    wb.save(out_path)

    print(f"対象: {run_dir}")
    if compare_dir:
        print(f"比較: {compare_dir}")
    print(f"出力: {out_path}")
    print(f"  シート: {', '.join(wb.sheetnames)}")
    print(f"  試行数 {len(results)} / 2mm未満 {within(results, 2.0)}")


if __name__ == "__main__":
    main()
