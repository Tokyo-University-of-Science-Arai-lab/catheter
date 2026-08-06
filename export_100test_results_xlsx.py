#!/usr/bin/env python3
"""
offline_POINTCLOUD_DEBUG_SAM3.py の results.csv / summary.json を Excel (.xlsx) に変換する。

出力: <run_dir>/results.xlsx
  ・summary シート: 全体の精度（しきい値別の割合・平均/中央値誤差）と品目別の内訳
  ・results シート: 試行ごとの詳細（display_name列を追加、誤差で色分け）

使い方:
  python export_100test_results_xlsx.py                # 最新の実行結果を変換
  python export_100test_results_xlsx.py <run_dir>       # 実行結果を指定
"""

import csv
import json
import statistics
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
OFFLINE_BASE_DIR = BASE_DIR / "captures" / "100test_offline"
MASTER_JSON = BASE_DIR / "master_100test.json"

# 誤差がこれ以内なら「良好」として色を付ける [mm]
GOOD_ERROR_MM = 2.0
# 誤差がこれを超えたら「大外れ」として色を付ける [mm]
BAD_ERROR_MM = 10.0

RESULT_COLUMNS = [
    ("test_index",          "試行",           True,   6),
    ("master_index",        "品目番号",       True,   9),
    ("repeat_index",        "周回",           True,   6),
    ("display_name",        "品名",           False, 28),
    ("book_name",           "型番/query",     False, 18),
    ("gt_book_width_mm",    "正解幅[mm]",     True,  11),
    ("pred_book_width_mm",  "推定幅[mm]",     True,  11),
    ("signed_error_mm",     "誤差[mm]",       True,  11),
    ("abs_error_mm",        "絶対誤差[mm]",   True,  12),
    ("elapsed_sec",         "所要[秒]",       True,   9),
    ("status",              "状態",           False,  9),
    ("error",               "エラー内容",     False, 30),
    ("run_shot_dir",        "出力フォルダ",   False, 60),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GOOD_FILL   = PatternFill("solid", fgColor="C6EFCE")   # 緑系: 良好（2mm以内）
BAD_FILL    = PatternFill("solid", fgColor="FFC7CE")   # 赤系: 大外れ（10mm超）
FAIL_FILL   = PatternFill("solid", fgColor="FFEB9C")   # 黄系: 認識失敗


def latest_run_dir(base_dir: Path) -> Path:
    runs = [p for p in base_dir.iterdir() if p.is_dir() and (p / "results.csv").exists()]
    if not runs:
        raise FileNotFoundError(f"results.csv を持つ実行結果がありません: {base_dir}")
    return max(runs, key=lambda p: p.stat().st_mtime)


def load_display_names(master_json: Path) -> dict[str, str]:
    if not master_json.exists():
        return {}
    data = json.loads(master_json.read_text(encoding="utf-8"))
    return {
        item["book_name"]: item.get("_display_name") or item["book_name"]
        for item in data
    }


def to_number(value):
    """数値に変換できればfloat/intで返す。できなければ元の文字列のまま返す。"""
    if value in (None, ""):
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return value
    return int(f) if f.is_integer() and abs(f) < 1e9 else f


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_summary_sheet(ws, rows, summary):
    thresholds = [1.0, 1.5, 2.0]
    total = summary.get("total_trials", len(rows))

    ws.append(["全体の精度"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["総試行数", total])
    ws.append(["成功", summary.get("success_count")])
    ws.append(["失敗", summary.get("fail_count")])
    ws.append([])
    ws.append(["しきい値", "件数", "割合"])
    style_header(ws, row=ws.max_row)
    for th in thresholds:
        key = f"within_{th:.1f}mm"
        tc = summary.get("threshold_counts", {}).get(key, {})
        ws.append([f"{th:.1f}mm未満", tc.get("count"), tc.get("rate")])
        ws.cell(row=ws.max_row, column=3).number_format = "0.0%"
    ws.append([])
    ws.append(["平均絶対誤差[mm]", round(summary.get("mean_abs_error_mm_success_only", 0) or 0, 3)])
    ws.append(["中央値絶対誤差[mm]", round(summary.get("median_abs_error_mm_success_only", 0) or 0, 3)])
    ws.append(["最大絶対誤差[mm]", round(summary.get("max_abs_error_mm_success_only", 0) or 0, 3)])
    ws.append(["最小絶対誤差[mm]", round(summary.get("min_abs_error_mm_success_only", 0) or 0, 3)])
    ws.append([])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append([
        "品目番号", "品名", "型番", "正解幅[mm]",
        "平均絶対誤差[mm]", "平均符号付誤差[mm]", "標準偏差[mm]",
        "最小[mm]", "最大[mm]", "2mm以内件数", "試行数",
    ])
    style_header(ws, row=header_row)

    by_item = defaultdict(list)
    for r in rows:
        by_item[(r.get("master_index"), r.get("book_name"))].append(r)

    for (master_index, book_name), item_rows in sorted(
        by_item.items(), key=lambda kv: int(kv[0][0]) if kv[0][0] not in (None, "") else 0
    ):
        errs = [to_number(r.get("abs_error_mm")) for r in item_rows if to_number(r.get("abs_error_mm")) is not None]
        signed = [to_number(r.get("signed_error_mm")) for r in item_rows if to_number(r.get("signed_error_mm")) is not None]
        gt = to_number(item_rows[0].get("gt_book_width_mm"))
        display_name = item_rows[0].get("display_name", book_name)
        good_count = sum(1 for e in errs if e <= GOOD_ERROR_MM)
        ws.append([
            to_number(master_index),
            display_name,
            book_name,
            gt,
            round(sum(errs) / len(errs), 3) if errs else None,
            round(sum(signed) / len(signed), 3) if signed else None,
            round(statistics.pstdev(signed), 3) if len(signed) > 1 else 0.0,
            round(min(errs), 3) if errs else None,
            round(max(errs), 3) if errs else None,
            good_count,
            len(item_rows),
        ])

    widths = [9, 26, 16, 11, 14, 14, 11, 9, 9, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row + 1}"


def build_results_sheet(ws, rows):
    ws.append([label for _, label, _, _ in RESULT_COLUMNS])
    style_header(ws)

    for r in rows:
        ws.append([
            to_number(r.get(key)) if numeric else r.get(key, "")
            for key, _, numeric, _ in RESULT_COLUMNS
        ])

    error_cols = {"signed_error_mm", "abs_error_mm", "pred_book_width_mm", "gt_book_width_mm", "elapsed_sec"}
    for i, (key, _, _, width) in enumerate(RESULT_COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = width
        if key in error_cols:
            for cell in ws[letter][1:]:
                cell.number_format = "0.000"

    abs_col_idx = next(i for i, (k, _, _, _) in enumerate(RESULT_COLUMNS, start=1) if k == "abs_error_mm")
    status_col_idx = next(i for i, (k, _, _, _) in enumerate(RESULT_COLUMNS, start=1) if k == "status")
    n_good = n_bad = n_fail = 0
    for row in ws.iter_rows(min_row=2):
        status = row[status_col_idx - 1].value
        err = row[abs_col_idx - 1].value

        if status != "success":
            fill, hit = FAIL_FILL, "fail"
        elif isinstance(err, (int, float)) and err > BAD_ERROR_MM:
            fill, hit = BAD_FILL, "bad"
        elif isinstance(err, (int, float)) and err <= GOOD_ERROR_MM:
            fill, hit = GOOD_FILL, "good"
        else:
            continue

        for cell in row:
            cell.fill = fill
        n_fail += hit == "fail"
        n_bad += hit == "bad"
        n_good += hit == "good"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return n_good, n_bad, n_fail


def main():
    run_dir = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else latest_run_dir(OFFLINE_BASE_DIR)
    results_csv = run_dir / "results.csv"
    summary_json = run_dir / "summary.json"
    if not results_csv.exists():
        raise FileNotFoundError(f"results.csv がありません: {results_csv}")

    display_names = load_display_names(MASTER_JSON)
    summary = json.loads(summary_json.read_text(encoding="utf-8")) if summary_json.exists() else {}

    with open(results_csv, encoding="utf-8-sig") as f:
        rows = [dict(r) for r in csv.DictReader(f)]
    for r in rows:
        r["display_name"] = display_names.get(r.get("book_name", ""), r.get("book_name", ""))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"
    build_summary_sheet(ws_summary, rows, summary)

    ws_results = wb.create_sheet("results")
    n_good, n_bad, n_fail = build_results_sheet(ws_results, rows)

    out_path = run_dir / "results.xlsx"
    wb.save(out_path)

    print(f"対象: {results_csv}")
    print(f"出力: {out_path}")
    print(f"  {len(rows)} 行")
    print(f"  緑（誤差{GOOD_ERROR_MM}mm以内）: {n_good}")
    print(f"  赤（誤差{BAD_ERROR_MM}mm超）: {n_bad}")
    print(f"  黄（認識失敗）: {n_fail}")


if __name__ == "__main__":
    main()
